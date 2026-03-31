from pathlib import Path

from openpyxl import load_workbook

from src.kd_detector import write_batch_summary_excel


def test_write_batch_summary_excel_contains_expected_headers(tmp_path: Path) -> None:
    rows = [
        {
            "image_name": "a.png",
            "lane_x_index": 1,
            "y_image": 200,
            "kD": 75.6,
            "gray_value": 132.4,
            "peak_height": 45.6,
            "prominence": 12.3,
        }
    ]
    out = tmp_path / "batch_summary.xlsx"
    write_batch_summary_excel(rows, out)

    wb = load_workbook(out)
    ws = wb["bands"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["image_name", "x", "y", "kD", "gray_mean", "peak_height", "prominence"]
    assert ws.cell(row=2, column=1).value == "a.png"
    assert ws.cell(row=2, column=2).value == 1
