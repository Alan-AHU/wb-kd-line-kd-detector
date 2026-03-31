from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from scipy.signal import find_peaks


KD_REFERENCES = np.array([250.0, 150.0, 100.0, 70.0, 50.0, 40.0, 35.0], dtype=float)
DEFAULT_AXIS_TICKS_557H = np.array([322.0, 360.0, 398.0, 452.0, 487.0, 512.0, 525.0], dtype=float)

SINGLE_EXPORT_FIELDS: list[tuple[str, str]] = [
    ("lane_x_index", "x"),
    ("y_image", "y"),
    ("kD", "kD"),
    ("gray_value", "gray_mean"),
    ("peak_height", "peak_height"),
    ("prominence", "prominence"),
]

BATCH_EXPORT_FIELDS: list[tuple[str, str]] = [("image_name", "image_name"), *SINGLE_EXPORT_FIELDS]


@dataclass(frozen=True)
class BlotRoi:
    x0: int
    x1: int
    y0: int
    y1: int

    @property
    def width(self) -> int:
        return self.x1 - self.x0

    @property
    def height(self) -> int:
        return self.y1 - self.y0


def _longest_true_run(mask: np.ndarray) -> tuple[int, int]:
    start = None
    best = (0, 0)
    for i, value in enumerate(mask):
        if value and start is None:
            start = i
        if (not value or i == len(mask) - 1) and start is not None:
            end = i if not value else i + 1
            if (end - start) > (best[1] - best[0]):
                best = (start, end)
            start = None
    return best


def detect_blot_roi(gray: np.ndarray) -> BlotRoi:
    h, w = gray.shape

    # 1) White-background suppression: keep only non-white foreground.
    white_thr = float(np.percentile(gray, 92))
    white_thr = min(250.0, max(232.0, white_thr))
    fg = (gray < white_thr).astype(np.uint8)

    # 2) Local density map: text strokes are sparse, blot panel is dense.
    density = cv2.boxFilter(fg.astype(np.float32), ddepth=-1, ksize=(31, 31), normalize=True)
    dense = (density > 0.14).astype(np.uint8)
    dense = cv2.morphologyEx(dense, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_RECT, (25, 25)), iterations=1)
    dense = cv2.morphologyEx(dense, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7)), iterations=1)

    best_bbox: tuple[int, int, int, int] | None = None
    best_score = -1.0
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(dense, connectivity=8)
    for i in range(1, num_labels):
        x, y, ww, hh, area = stats[i]
        if area < int(0.015 * h * w):
            continue
        if ww < int(0.18 * w) or hh < int(0.20 * h):
            continue
        if (y + hh) < int(0.62 * h):
            continue
        cx = x + ww / 2.0
        if cx > 0.82 * w:
            continue
        if x > int(0.45 * w):
            continue
        ar = ww / (hh + 1e-6)
        if not (0.30 <= ar <= 1.90):
            continue
        left_bonus = max(0.0, 0.80 - (cx / max(1.0, float(w))))
        score = float(area) * (1.0 + 0.6 * left_bonus)
        if score > best_score:
            best_score = score
            best_bbox = (int(x), int(y), int(ww), int(hh))

    if best_bbox is not None:
        x, y, ww, hh = best_bbox
        pad_x = max(8, int(0.04 * ww))
        pad_y = max(6, int(0.04 * hh))
        x0 = max(0, x - pad_x)
        x1 = min(w, x + ww + pad_x)
        y0 = max(0, y - pad_y)
        y1 = min(h, y + hh + pad_y)
        if (x1 - x0) >= int(0.20 * w) and (y1 - y0) >= int(0.25 * h):
            return BlotRoi(x0=x0, x1=x1, y0=y0, y1=y1)

    # Fallback if foreground-box extraction fails on edge cases.
    x0, x1 = int(0.08 * w), int(0.62 * w)
    y0, y1 = int(0.42 * h), int(0.98 * h)
    return BlotRoi(x0=x0, x1=x1, y0=y0, y1=y1)


def detect_lane_centers(roi_gray: np.ndarray) -> list[int]:
    width = roi_gray.shape[1]
    inv = 255.0 - roi_gray.astype(np.float32)
    mean_profile = inv.mean(axis=0)
    dark_thr = float(np.percentile(roi_gray, 45))
    dark_ratio_profile = (roi_gray <= dark_thr).mean(axis=0) * 90.0
    profile = 0.70 * mean_profile + 0.30 * dark_ratio_profile
    smoothed = np.convolve(profile, np.ones(9) / 9, mode="same")

    long_win = max(31, int(0.18 * width))
    if long_win % 2 == 0:
        long_win += 1
    baseline = np.convolve(smoothed, np.ones(long_win) / float(long_win), mode="same")
    signal = smoothed - baseline

    min_distance = max(8, int(0.045 * width))
    p90_smoothed = float(np.percentile(smoothed, 90))
    p65_smoothed = float(np.percentile(smoothed, 65))
    p90 = float(np.percentile(signal, 90))
    p70 = float(np.percentile(signal, 70))
    peaks, props = find_peaks(
        smoothed,
        distance=min_distance,
        prominence=max(0.5, p90_smoothed * 0.035),
        height=max(0.5, p65_smoothed),
        width=1.0,
    )

    if len(peaks) < 2:
        peaks, props = find_peaks(
            signal,
            distance=max(8, int(0.045 * width)),
            prominence=max(0.3, p90 * 0.06),
            height=max(0.2, float(np.percentile(signal, 58))),
            width=0.8,
        )

    candidates: list[tuple[int, float]] = []
    prominences = props.get("prominences", np.zeros(len(peaks), dtype=float))
    heights = props.get("peak_heights", np.zeros(len(peaks), dtype=float))
    for idx, p in enumerate(peaks.astype(int).tolist()):
        if p < 3 or p > (width - 4):
            continue
        score = float(prominences[idx] * 1.7 + heights[idx])
        candidates.append((p, score))

    if not candidates:
        return []

    candidates.sort(key=lambda t: t[0])
    merged: list[tuple[int, float]] = []
    near_gap = max(5, int(0.022 * width))
    for p, score in candidates:
        if not merged or (p - merged[-1][0]) > near_gap:
            merged.append((p, score))
        elif score > merged[-1][1]:
            merged[-1] = (p, score)

    max_lanes = 8
    if len(merged) > max_lanes:
        merged = sorted(sorted(merged, key=lambda t: t[1], reverse=True)[:max_lanes], key=lambda t: t[0])

    lane_centers = [int(p) for p, _ in merged]
    lane_centers.sort()

    if len(lane_centers) >= 2 and len(lane_centers) < max_lanes:
        step = float(np.median(np.diff(lane_centers)))
        if step > 6:
            filled = [lane_centers[0]]
            for p in lane_centers[1:]:
                prev = filled[-1]
                gap = p - prev
                if gap > (1.8 * step) and len(filled) < max_lanes:
                    filled.append(int(round((prev + p) / 2.0)))
                filled.append(p)
            lane_centers = sorted(filled)[:max_lanes]

    return lane_centers


def _build_axis_tick_signal(gray: np.ndarray, x0: int, x1: int) -> np.ndarray:
    strip = gray[:, x0:x1]
    if strip.size == 0:
        return np.zeros(gray.shape[0], dtype=float)

    # Axis ticks are short dark horizontal segments; use a conservative dark threshold.
    dark_thr = min(245.0, float(np.percentile(strip, 70)))
    binary = strip <= dark_thr

    row_counts = binary.sum(axis=1).astype(float)
    run_lengths = np.zeros(strip.shape[0], dtype=float)
    for yi, row in enumerate(binary):
        cur = 0
        mx = 0
        for v in row:
            if v:
                cur += 1
                if cur > mx:
                    mx = cur
            else:
                cur = 0
        run_lengths[yi] = float(mx)

    signal = 0.70 * run_lengths + 0.30 * row_counts
    signal = np.convolve(signal, np.ones(5) / 5, mode="same")
    return signal


def _align_ticks_to_template(candidates: list[int], template: list[int], tolerance: int) -> list[int]:
    if not candidates:
        return template

    used: set[int] = set()
    out: list[int] = []
    for t in template:
        best_idx = -1
        best_dist = 10**9
        for i, c in enumerate(candidates):
            if i in used:
                continue
            d = abs(int(c) - int(t))
            if d < best_dist:
                best_dist = d
                best_idx = i
        if best_idx >= 0 and best_dist <= tolerance:
            used.add(best_idx)
            out.append(int(candidates[best_idx]))
        else:
            out.append(int(t))

    out = sorted(out)
    return out


def detect_scale_ticks(gray: np.ndarray, roi: BlotRoi, expected_count: int = 7) -> list[int]:
    expected_count = max(2, min(expected_count, len(KD_REFERENCES)))
    h, w = gray.shape

    default_scaled = np.clip(np.round(DEFAULT_AXIS_TICKS_557H / 557.0 * h), 0, h - 1).astype(int).tolist()
    template = default_scaled[:expected_count]

    x0 = max(0, roi.x0 - 55)
    x1 = max(x0 + 18, roi.x0 - 4)
    signal = _build_axis_tick_signal(gray, x0=x0, x1=x1)

    ymin = max(0, roi.y0 - 10)
    ymax = min(h - 1, roi.y1 + 8)
    segment = signal[ymin:ymax]
    if len(segment) < 8:
        return template

    dist = max(8, int(0.014 * h))
    base_height = float(np.percentile(segment, 82))
    peaks, props = find_peaks(segment, distance=dist, prominence=0.8, height=base_height)
    if len(peaks) < 4:
        peaks, props = find_peaks(
            segment,
            distance=max(6, dist - 2),
            prominence=0.4,
            height=float(np.percentile(segment, 72)),
        )

    candidates = sorted((peaks + ymin).astype(int).tolist())
    if not candidates:
        return template

    tolerance = max(10, int(0.03 * h))
    aligned = _align_ticks_to_template(candidates, template, tolerance=tolerance)
    return aligned


def fit_log_kd_mapping(scale_y: list[int], kd_refs: np.ndarray | None = None) -> tuple[float, float]:
    refs = KD_REFERENCES if kd_refs is None else kd_refs
    y = np.array(scale_y, dtype=float)
    if len(y) != len(refs):
        count = min(len(y), len(refs))
        y = y[:count]
        refs = refs[:count]

    a, b = np.linalg.lstsq(np.vstack([y, np.ones_like(y)]).T, np.log10(refs), rcond=None)[0]
    return float(a), float(b)


def y_to_kd(y: float, a: float, b: float) -> float:
    return float(10 ** (a * y + b))


def make_piecewise_kd_mapper(scale_y: list[int], kd_refs: np.ndarray | None = None) -> Callable[[float], float]:
    refs = KD_REFERENCES if kd_refs is None else np.array(kd_refs, dtype=float)
    y = np.array(scale_y, dtype=float)
    if len(y) != len(refs):
        count = min(len(y), len(refs))
        y = y[:count]
        refs = refs[:count]

    order = np.argsort(y)
    y = y[order]
    kd_vals = refs[order]

    # Fallback to global linear mapping if scale points are insufficient.
    if len(y) < 2:
        a, b = fit_log_kd_mapping(scale_y=list(y.astype(float)), kd_refs=refs)
        return lambda yy: y_to_kd(yy, a, b)

    left_slope = (kd_vals[1] - kd_vals[0]) / (y[1] - y[0] + 1e-9)
    right_slope = (kd_vals[-1] - kd_vals[-2]) / (y[-1] - y[-2] + 1e-9)

    def _mapper(yy: float) -> float:
        yy_f = float(yy)
        if yy_f < y[0]:
            return float(kd_vals[0] + left_slope * (yy_f - y[0]))
        if yy_f > y[-1]:
            return float(kd_vals[-1] + right_slope * (yy_f - y[-1]))
        return float(np.interp(yy_f, y, kd_vals))

    return _mapper


def _compute_band_gray_metrics(lane_strip: np.ndarray, y_center: int) -> dict[str, float]:
    lane_h = lane_strip.shape[0]
    y0 = max(0, int(y_center) - 2)
    y1 = min(lane_h, int(y_center) + 3)
    band_patch = lane_strip[y0:y1, :].astype(float)
    if band_patch.size == 0:
        lane_mean = float(lane_strip.mean()) if lane_strip.size > 0 else 255.0
        return {
            "gray_value": lane_mean,
            "gray_min": lane_mean,
            "darkness": 255.0 - lane_mean,
            "local_contrast": 0.0,
        }

    gray_value = float(band_patch.mean())
    gray_min = float(np.percentile(band_patch, 10))
    darkness = 255.0 - gray_value

    bg_parts: list[np.ndarray] = []
    up0 = max(0, y0 - 10)
    up1 = max(0, y0 - 4)
    dn0 = min(lane_h, y1 + 4)
    dn1 = min(lane_h, y1 + 10)
    if up1 > up0:
        bg_parts.append(lane_strip[up0:up1, :].astype(float))
    if dn1 > dn0:
        bg_parts.append(lane_strip[dn0:dn1, :].astype(float))

    if bg_parts:
        bg_values = np.concatenate([part.reshape(-1) for part in bg_parts])
        bg_mean = float(bg_values.mean())
    else:
        bg_mean = float(lane_strip.mean())
    local_contrast = bg_mean - gray_value

    y_safe = min(max(int(y_center), 0), lane_h - 1) if lane_h > 0 else 0
    center_row = lane_strip[y_safe, :] if lane_h > 0 else np.zeros((0,), dtype=float)
    if center_row.size > 0:
        dark_thr = min(float(np.percentile(lane_strip, 40)), float(np.percentile(center_row, 70)))
        dark_mask = center_row <= dark_thr
        max_run = 0
        cur_run = 0
        for v in dark_mask:
            if v:
                cur_run += 1
                if cur_run > max_run:
                    max_run = cur_run
            else:
                cur_run = 0
        dark_run = float(max_run)
    else:
        dark_run = 0.0

    return {
        "gray_value": gray_value,
        "gray_min": gray_min,
        "darkness": darkness,
        "local_contrast": local_contrast,
        "dark_run": dark_run,
    }


def detect_bands_by_lane(
    roi_gray: np.ndarray,
    lane_centers: list[int],
    roi_top_y: int,
    roi_left_x: int,
    kd_from_y: Callable[[float], float],
    scale_y: list[int],
    axis_x_range: tuple[int, int] | None = None,
    axis_y_range_abs: tuple[int, int] | None = None,
) -> list[dict[str, Any]]:
    min_scale_y = min(scale_y)
    max_scale_y = max(scale_y)
    results: list[dict[str, Any]] = []

    roi_h, roi_w = roi_gray.shape
    x_min = 0
    x_max = roi_w - 1
    if axis_x_range is not None:
        x_min = max(0, int(axis_x_range[0]))
        x_max = min(roi_w - 1, int(axis_x_range[1]))

    y_local_min = 0
    y_local_max = roi_h - 1
    if axis_y_range_abs is not None:
        y_abs_min = int(axis_y_range_abs[0])
        y_abs_max = int(axis_y_range_abs[1])
        y_local_min = max(0, y_abs_min - roi_top_y)
        y_local_max = min(roi_h - 1, y_abs_max - roi_top_y)
        if y_local_min > y_local_max:
            y_local_min, y_local_max = 0, roi_h - 1

    filtered_centers = [c for c in lane_centers if x_min <= int(c) <= x_max]
    if filtered_centers:
        lane_centers = filtered_centers

    if len(lane_centers) >= 2:
        lane_step = float(np.median(np.diff(sorted(int(v) for v in lane_centers))))
    else:
        lane_step = 18.0
    half_lane_w = int(min(12, max(8, round(0.34 * lane_step))))

    for lane_idx, center_x in enumerate(lane_centers):
        left = max(0, center_x - half_lane_w)
        right = min(roi_gray.shape[1], center_x + half_lane_w + 1)
        lane_strip = roi_gray[y_local_min : y_local_max + 1, left:right]
        if lane_strip.size == 0:
            continue

        lane_strip_f = lane_strip.astype(np.float32)
        lane_blur = cv2.GaussianBlur(lane_strip_f, (3, 3), 0)

        # Local thresholding (per-lane) to capture dark stripe pixels.
        pix_thr = float(np.percentile(lane_blur, 34))
        pix_thr = min(pix_thr, float(np.percentile(lane_blur, 48)) - 2.0)
        pix_thr = max(25.0, min(230.0, pix_thr))
        dark_mask = (lane_blur <= pix_thr).astype(np.uint8)
        dark_mask = cv2.morphologyEx(
            dark_mask,
            cv2.MORPH_OPEN,
            cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2)),
            iterations=1,
        )
        dark_mask = cv2.morphologyEx(
            dark_mask,
            cv2.MORPH_CLOSE,
            cv2.getStructuringElement(cv2.MORPH_RECT, (3, 2)),
            iterations=1,
        )

        row_ratio = dark_mask.mean(axis=1).astype(float)
        row_darkness = (255.0 - lane_blur).mean(axis=1).astype(float)
        row_dark_p30 = (255.0 - np.percentile(lane_blur, 30, axis=1)).astype(float)
        row_score = 0.65 * row_darkness + 0.35 * row_dark_p30
        baseline = cv2.GaussianBlur(row_score.reshape(-1, 1), (1, 41), 0).ravel()
        row_contrast = row_score - baseline

        # Two-level gating: strict rows + relaxed rows near local maxima.
        ratio_thr = max(0.13, float(np.percentile(row_ratio, 74)) * 0.50)
        contrast_thr = max(0.55, float(np.percentile(row_contrast, 70)) * 0.48)
        score_thr = max(7.0, float(np.percentile(row_score, 68)) * 0.62)

        ratio_thr_relaxed = max(0.08, ratio_thr * 0.72)
        contrast_thr_relaxed = max(0.30, contrast_thr * 0.68)
        score_thr_relaxed = max(4.8, score_thr * 0.70)

        strict_rows = (row_ratio >= ratio_thr) & ((row_contrast >= contrast_thr) | (row_score >= score_thr))
        relaxed_rows = (row_ratio >= ratio_thr_relaxed) & (
            (row_contrast >= contrast_thr_relaxed) | (row_score >= score_thr_relaxed)
        )

        score_smooth = np.convolve(row_score, np.ones(5) / 5.0, mode="same")
        contrast_pos = np.maximum(row_contrast, 0.0)
        peaks_score, _ = find_peaks(
            score_smooth,
            distance=2,
            prominence=max(0.10, float(np.percentile(score_smooth, 70)) * 0.02),
        )
        peaks_contrast, _ = find_peaks(
            contrast_pos,
            distance=2,
            prominence=max(0.08, float(np.percentile(contrast_pos, 72)) * 0.04),
        )
        peak_mask = np.zeros_like(relaxed_rows, dtype=bool)
        for pk in np.concatenate([peaks_score, peaks_contrast]).astype(int).tolist():
            p0 = max(0, pk - 1)
            p1 = min(len(peak_mask), pk + 2)
            peak_mask[p0:p1] = True

        row_candidate = strict_rows | (relaxed_rows & peak_mask)
        row_candidate_u8 = row_candidate.astype(np.uint8)
        if len(row_candidate_u8) >= 5:
            row_candidate_u8 = (
                np.convolve(row_candidate_u8, np.ones(5, dtype=np.uint8), mode="same") >= 2
            ).astype(np.uint8)

        runs: list[tuple[int, int]] = []
        start = None
        for yi, flag in enumerate(row_candidate_u8.tolist()):
            if flag and start is None:
                start = yi
            if (not flag or yi == len(row_candidate_u8) - 1) and start is not None:
                end = yi if not flag else yi + 1
                if (end - start) >= 2:
                    runs.append((int(start), int(end)))
                start = None

        lane_x_index = lane_idx + 1
        lane_label = str(lane_x_index)
        lane_rows: list[dict[str, Any]] = []
        for y0_run, y1_run in runs:
            local_scores = row_score[y0_run:y1_run] + 0.7 * np.maximum(row_contrast[y0_run:y1_run], 0.0)
            if len(local_scores) == 0:
                continue
            y_refined = int(y0_run + int(np.argmax(local_scores)))

            metrics = _compute_band_gray_metrics(lane_blur, y_refined)
            darkness_floor = max(5.0, float(np.percentile(row_darkness, 45)) * 0.42)
            if metrics["dark_run"] < 2.0 and row_ratio[y_refined] < (ratio_thr_relaxed * 1.05):
                continue
            if metrics["local_contrast"] < 0.25 and row_contrast[y_refined] < (contrast_thr_relaxed * 0.90):
                continue
            if metrics["darkness"] < darkness_floor and row_score[y_refined] < score_thr_relaxed:
                continue

            y_abs = int(roi_top_y + y_local_min + y_refined)
            if y_abs < (min_scale_y - 10) or y_abs > (max_scale_y + 10):
                continue

            kd = float(kd_from_y(y_abs))
            extrapolated = y_abs < min_scale_y or y_abs > max_scale_y
            x_abs = int(roi_left_x + center_x)
            prominence = float(max(0.0, row_contrast[y_refined]))
            peak_height = float(max(0.0, row_score[y_refined]))

            lane_rows.append(
                {
                    "lane_index": lane_idx,
                    "lane_x_index": lane_x_index,
                    "lane_label": lane_label,
                    "x": int(center_x),
                    "x_image": x_abs,
                    "y": y_abs,
                    "y_image": y_abs,
                    "y_roi": int(y_refined),
                    "kD": kd,
                    "gray_value": float(metrics["gray_value"]),
                    "gray_min": float(metrics["gray_min"]),
                    "darkness": float(metrics["darkness"]),
                    "local_contrast": float(metrics["local_contrast"]),
                    "prominence": prominence,
                    "peak_height": peak_height,
                    "extrapolated": extrapolated,
                }
            )

        lane_rows.sort(key=lambda r: int(r["y"]))
        deduped_rows: list[dict[str, Any]] = []
        for row in lane_rows:
            if not deduped_rows or abs(int(row["y"]) - int(deduped_rows[-1]["y"])) > 3:
                deduped_rows.append(row)
                continue
            if float(row["darkness"]) > float(deduped_rows[-1]["darkness"]):
                deduped_rows[-1] = row

        for band_idx, row in enumerate(deduped_rows, start=1):
            row["band_index_in_lane"] = band_idx

        results.extend(deduped_rows)

    return results


def draw_overlay(
    image: np.ndarray,
    roi: BlotRoi,
    lane_centers: list[int],
    bands: list[dict[str, Any]],
    output_path: Path,
) -> None:
    vis = image.copy()
    cv2.rectangle(vis, (roi.x0, roi.y0), (roi.x1, roi.y1), (0, 255, 0), 2)

    lane_label_y = roi.y0 - 10 if roi.y0 > 24 else min(vis.shape[0] - 8, roi.y0 + 16)
    for lane_idx, cx in enumerate(lane_centers, start=1):
        x_abs = roi.x0 + cx
        cv2.line(vis, (x_abs, roi.y0), (x_abs, roi.y1), (255, 0, 0), 1)
        cv2.putText(
            vis,
            str(lane_idx),
            (x_abs - 8, lane_label_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.42,
            (255, 0, 0),
            1,
            cv2.LINE_AA,
        )

    # Smaller font + simple per-lane vertical nudging to reduce overlap.
    label_font = cv2.FONT_HERSHEY_SIMPLEX
    label_scale = 0.28
    label_thickness = 1
    placed_y_by_lane: dict[int, list[int]] = {}

    for band in sorted(bands, key=lambda r: (int(r.get("lane_x_index", 1)), int(r["y"]))):
        x_abs = roi.x0 + int(band["x"])
        y_abs = int(band["y"])
        cv2.circle(vis, (x_abs, y_abs), 3, (0, 0, 255), -1)

        text = f"{band['kD']:.1f}kD"
        (text_w, text_h), _ = cv2.getTextSize(text, label_font, label_scale, label_thickness)
        tx = min(vis.shape[1] - text_w - 2, x_abs + 3)
        ty = max(text_h + 2, y_abs - 2)

        lane_id = int(band.get("lane_x_index", 1))
        lane_ys = placed_y_by_lane.setdefault(lane_id, [])
        min_gap = text_h + 2
        while any(abs(ty - old_ty) < min_gap for old_ty in lane_ys):
            ty = min(vis.shape[0] - 2, ty + min_gap)
            if ty >= vis.shape[0] - 2:
                break
        lane_ys.append(ty)

        cv2.putText(
            vis,
            text,
            (tx, ty),
            label_font,
            label_scale,
            (0, 0, 255),
            label_thickness,
            cv2.LINE_AA,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(str(output_path), vis)


def _format_sheet_for_readability(ws: Any) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(28, max(10, max_len + 2))


ROUND_2_KEYS = {"kD", "gray_value", "peak_height", "prominence"}


def _extract_export_value(row: dict[str, Any], key: str, image_name: str | None = None) -> Any:
    if key == "image_name":
        return image_name
    return row[key]


def _format_export_value_for_csv(key: str, value: Any) -> Any:
    if key in ROUND_2_KEYS:
        return f"{float(value):.2f}"
    return value


def write_outputs(
    output_dir: Path,
    image_name: str,
    roi: BlotRoi,
    lane_centers: list[int],
    scale_y: list[int],
    a: float,
    b: float,
    bands: list[dict[str, Any]],
) -> tuple[Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"{image_name}_bands.csv"
    json_path = output_dir / f"{image_name}_bands.json"
    xlsx_path = output_dir / f"{image_name}_bands.xlsx"
    single_headers = [header for _, header in SINGLE_EXPORT_FIELDS]

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=single_headers)
        writer.writeheader()
        for row in bands:
            out_row: dict[str, Any] = {}
            for key, header in SINGLE_EXPORT_FIELDS:
                value = _extract_export_value(row, key)
                out_row[header] = _format_export_value_for_csv(key, value)
            writer.writerow(out_row)

    summary = {
        "roi": {"x0": roi.x0, "x1": roi.x1, "y0": roi.y0, "y1": roi.y1},
        "lane_centers": lane_centers,
        "scale_y": scale_y,
        "mapping": {
            "mode": "piecewise_interpolation",
            "piecewise_points": [{"y": int(y), "kD": float(kd)} for y, kd in zip(scale_y, KD_REFERENCES[: len(scale_y)])],
            "linear_reference": {"log10_kD = a*y + b": {"a": a, "b": b}},
        },
        "band_measurement": {
            "gray_mean": "mean grayscale in a 5-row window centered at the band (0=black, 255=white)",
            "peak_height": "height of the detected peak in lane signal",
            "prominence": "prominence of the detected peak in lane signal",
        },
        "bands": bands,
    }
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "bands"
    ws.append([header for _, header in SINGLE_EXPORT_FIELDS])
    for row in bands:
        values: list[Any] = []
        for key, _header in SINGLE_EXPORT_FIELDS:
            value = _extract_export_value(row, key)
            if key in ROUND_2_KEYS:
                value = round(float(value), 2)
            values.append(value)
        ws.append(values)

    meta = wb.create_sheet("meta")
    meta.append(["parameter", "value"])
    meta.append(["roi_x0", roi.x0])
    meta.append(["roi_x1", roi.x1])
    meta.append(["roi_y0", roi.y0])
    meta.append(["roi_y1", roi.y1])
    meta.append(["scale_y", ",".join(str(v) for v in scale_y)])
    meta.append(["scale_kd_refs", ",".join(str(float(v)) for v in KD_REFERENCES[: len(scale_y)])])
    meta.append(["kd_mapping_mode", "piecewise_linear_interpolation"])
    meta.append(["reference_log_fit_a", a])
    meta.append(["reference_log_fit_b", b])
    meta.append(["lane_count", len(lane_centers)])
    meta.append(["gray_mean_method", "mean grayscale in a 5-row local patch around each band center"])

    _format_sheet_for_readability(ws)
    _format_sheet_for_readability(meta)
    wb.save(xlsx_path)

    return csv_path, json_path, xlsx_path


def write_batch_summary_excel(rows: list[dict[str, Any]], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "bands"
    ws.append([header for _, header in BATCH_EXPORT_FIELDS])

    for row in rows:
        values: list[Any] = []
        for key, _header in BATCH_EXPORT_FIELDS:
            value = _extract_export_value(row, key, image_name=str(row["image_name"]))
            if key in ROUND_2_KEYS:
                value = round(float(value), 2)
            values.append(value)
        ws.append(values)

    _format_sheet_for_readability(ws)
    wb.save(output_path)
    return output_path


def estimate_axis_box_ranges(
    roi_gray: np.ndarray,
    lane_centers: list[int],
    roi_top_y: int,
    scale_y: list[int],
) -> tuple[tuple[int, int], tuple[int, int]]:
    roi_h, roi_w = roi_gray.shape
    if lane_centers:
        centers = sorted(int(v) for v in lane_centers)
        if len(centers) >= 2:
            step = float(np.median(np.diff(centers)))
        else:
            step = float(max(12, int(0.06 * roi_w)))
        pad = max(8, int(round(0.75 * step)))
        x0 = max(0, int(centers[0] - pad))
        x1 = min(roi_w - 1, int(centers[-1] + pad))
    else:
        x0, x1 = 0, roi_w - 1

    # Keep detection strictly in the left-side axis box and avoid right text block.
    x1 = min(x1, int(0.90 * roi_w))

    y_abs_min = max(int(roi_top_y), int(min(scale_y) - 16))
    y_abs_max = min(int(roi_top_y + roi_h - 1), int(max(scale_y) + 16))
    if y_abs_min > y_abs_max:
        y_abs_min, y_abs_max = int(roi_top_y), int(roi_top_y + roi_h - 1)

    return (x0, x1), (y_abs_min, y_abs_max)


def analyze_image(
    image_path: Path,
    output_dir: Path,
    include_marker_lane: bool = True,
) -> dict[str, Any]:
    image = cv2.imread(str(image_path))
    if image is None:
        raise FileNotFoundError(f"Cannot read image: {image_path}")
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    roi = detect_blot_roi(gray)
    roi_gray = gray[roi.y0 : roi.y1, roi.x0 : roi.x1]
    lane_centers = detect_lane_centers(roi_gray)
    scale_y = detect_scale_ticks(gray, roi, expected_count=len(KD_REFERENCES))
    a, b = fit_log_kd_mapping(scale_y)
    kd_from_y = make_piecewise_kd_mapper(scale_y)
    axis_x_range, axis_y_range_abs = estimate_axis_box_ranges(roi_gray, lane_centers, roi.y0, scale_y)
    lane_centers_box = [c for c in lane_centers if axis_x_range[0] <= int(c) <= axis_x_range[1]]
    if lane_centers_box:
        lane_centers = lane_centers_box

    bands = detect_bands_by_lane(
        roi_gray,
        lane_centers,
        roi.y0,
        roi.x0,
        kd_from_y,
        scale_y,
        axis_x_range=axis_x_range,
        axis_y_range_abs=axis_y_range_abs,
    )
    if not include_marker_lane:
        bands = [bnd for bnd in bands if bnd["lane_index"] != 0]

    image_name = image_path.stem
    overlay_path = output_dir / f"{image_name}_annotated.png"
    draw_overlay(image, roi, lane_centers, bands, overlay_path)
    csv_path, json_path, xlsx_path = write_outputs(output_dir, image_name, roi, lane_centers, scale_y, a, b, bands)

    return {
        "image_path": str(image_path),
        "overlay_path": str(overlay_path),
        "csv_path": str(csv_path),
        "json_path": str(json_path),
        "xlsx_path": str(xlsx_path),
        "roi": roi,
        "lane_centers": lane_centers,
        "scale_y": scale_y,
        "mapping": {"mode": "piecewise", "linear_reference": {"a": a, "b": b}},
        "bands": bands,
    }

